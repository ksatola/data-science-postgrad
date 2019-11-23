from openpyxl import load_workbook
from openpyxl import Workbook
import os

# -----------------
# Set up
# Retrieve current working directory (`cwd`)
cwd = os.getcwd()
cwd

# Change directory 
os.chdir("./exercises")

# List all files and directories in current directory
os.listdir('.')

# -----------------
# Reading data from XLSX files

# open an existing Workbook (file)

wb = load_workbook('./data/01-basics.xlsx')

# print names of sheets
print(wb.sheetnames)


# get an active sheet
ws = wb.active

# print out a value in the first cell
a1 = ws['A1']
print(a1.value)

a2 = ws.cell(row=2, column=1)
print(a2.value)

# access a range of cells
cell_range = ws['A1:A10']
for col in cell_range:
    for cell in col:
        print(cell.value)

# access cols
colC = ws['A']
col_range = ws['C:D']

# access rows
row2 = ws[2]
row_range = ws[5:10]


# access cell values



# exercise #1.1 - avarage
sum = 0
cnt = 0
cell_range = ws['A1:A10']
for col in cell_range:
    for cell in col:
        sum += cell.value
        cnt += 1

print(sum / cnt)


# exercise #1.2 - the most frequent value
print('Exercise 1.2')
values = {}
cell_range = ws['A1:A10']
for col in cell_range:
    for cell in col:
        if cell.value in values:
            values[cell.value] = values[cell.value] + 1
        else: 
            values[cell.value] = 1

print(values)
        

# exercise #1.3 - prime numbers


# -----------------
# Writing data to XLSX files

wb = Workbook()

# A workbook is always created with at least one worksheet. You can get it by using the openpyxl.workbook.Workbook.active() property
ws = wb.active

# You can create a new sheet
ws1 = wb.create_sheet("Custom sheet")
ws1.title = "New title"
ws1.sheet_properties.tabColor = "1072BA"

# assign value to a cell
ws = wb.active
ws['A1'].value = 'hello world'

# assign values to a range
for row in ws.iter_rows(min_row=1, min_col=2, max_col=3, max_row=3):
    for cell in row:
        cell.value = 1

for i in range(5, 10):
    for j in range(5, 10):
        ws.cell(row=i, column=j).value = i * j

# saves workbook wb to file; the file is overwritten
wb.save('./data/01-tmp.xlsx')


# exercise #2.1 - generate random names

